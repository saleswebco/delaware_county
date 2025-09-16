# scrape_delaware.py
import asyncio
import json
import os
import traceback
from datetime import datetime
from pathlib import Path
import time

from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError

OUT_DIR = Path("out")
OUT_DIR.mkdir(exist_ok=True)


class DelawareScraper:
    def __init__(self, page, browser=None, context=None,
                 base_url: str = "https://delcorowonlineservices.co.delaware.pa.us/countyweb/loginDisplay.action?countyname=DelawarePA"):
        self.page = page
        self.browser = browser
        self.context = context
        self.base_url = base_url

    async def _dump_debug(self, name_prefix: str):
        """Save screenshot and page HTML for debugging."""
        ts = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
        shot = OUT_DIR / f"{name_prefix}-{ts}.png"
        html = OUT_DIR / f"{name_prefix}-{ts}.html"
        try:
            await self.page.screenshot(path=str(shot), full_page=True)
        except Exception as e:
            print("Failed to take screenshot:", e)
        try:
            html_content = await self.page.content()
            html.write_text(html_content, encoding="utf-8")
        except Exception as e:
            print("Failed to save HTML:", e)
        print(f"🧾 Debug dumped: {shot} , {html}")

    async def wait_for_frame_by_url_fragment(self, url_fragment: str, timeout: int = 60):
        """
        Poll the page.frames until a frame whose URL contains url_fragment appears.
        Returns the Frame object or raises TimeoutError.
        """
        print(f"⏳ Waiting for frame with url containing '{url_fragment}' (timeout {timeout}s)...")
        for i in range(timeout):
            for f in self.page.frames:
                if f.url and url_fragment in f.url:
                    print(f"✅ Found frame with url {f.url}")
                    return f
            await asyncio.sleep(1)
        raise PlaywrightTimeoutError(f"Frame with url fragment '{url_fragment}' not found within {timeout}s")

    # -----------------------
    # Login / navigation
    # -----------------------
    async def goto_login(self, retries: int = 3):
        """Go to login page and click 'Login as Guest'."""
        for attempt in range(1, retries + 1):
            try:
                await self.page.goto(self.base_url, timeout=60000)
                # small wait for initial rendering
                await self.page.wait_for_timeout(1500)
                # Try robust selector variations
                possible_selectors = [
                    "input[value=' Login as Guest ']",
                    "input[value='Login as Guest']",
                    "input[type='button'][value*='Guest']",
                    "text='Login as Guest'"
                ]
                clicked = False
                for sel in possible_selectors:
                    try:
                        await self.page.wait_for_selector(sel, timeout=5000)
                        await self.page.locator(sel).click()
                        clicked = True
                        break
                    except PlaywrightTimeoutError:
                        continue
                if not clicked:
                    raise PlaywrightTimeoutError("Could not find 'Login as Guest' button with known selectors")
                # wait for the expected post-login url pattern
                await self.page.wait_for_url("**/main.jsp?countyname=DelawarePA", timeout=60000)
                print("✅ Logged in as Guest")
                return
            except Exception as e:
                print(f"⚠ goto_login attempt {attempt} failed: {e}")
                if attempt == retries:
                    await self._dump_debug("goto_login_failed")
                    raise
                await asyncio.sleep(3)

    # -----------------------
    # Accept terms (iframe)
    # -----------------------
    async def accept_terms(self, retries: int = 3):
        """Handle iframe and click Accept button using frame_locator."""
        for attempt in range(1, retries + 1):
            try:
                # Wait for iframe wrapper presence (name might differ)
                await asyncio.sleep(1)
                # Use frame_locator to target the inner iframe's accept button.
                # Try several iframe selectors to be robust.
                iframe_selectors = [
                    "iframe[name='bodyframe']",
                    "iframe#bodyframe",
                    "iframe[src*='blank.jsp']",
                    "iframe",
                ]
                clicked = False
                for ifsel in iframe_selectors:
                    try:
                        frame_locator = self.page.frame_locator(ifsel)
                        accept = frame_locator.locator("#accept")
                        await accept.wait_for(state="visible", timeout=10000)
                        await accept.click()
                        clicked = True
                        break
                    except PlaywrightTimeoutError:
                        continue

                if not clicked:
                    # As a fallback: iterate frames and try to find an '#accept' element inside any frame
                    for f in self.page.frames:
                        try:
                            handle = await f.query_selector("#accept")
                            if handle:
                                await handle.click()
                                clicked = True
                                break
                        except Exception:
                            continue

                if not clicked:
                    raise PlaywrightTimeoutError("Could not locate an Accept button in known frame locations")

                print(f"✅ Accepted terms (attempt {attempt})")
                return

            except Exception as e:
                print(f"⚠ accept_terms attempt {attempt} failed: {e}")
                if attempt == retries:
                    print("Available frames (debug):")
                    for f in self.page.frames:
                        print(" -", f.url)
                    await self._dump_debug("accept_terms_failed")
                    raise
                await asyncio.sleep(2)

    # -----------------------
    # Click Search Public Records row
    # -----------------------
    async def click_search_public_records(self, retries: int = 3):
        """
        Click the row that triggers Search Public Records.
        Uses iframe frame_locator and robust waiting.
        """
        for attempt in range(1, retries + 1):
            try:
                await self.page.wait_for_timeout(1000)
                # The UI might place the table inside the same 'bodyframe' iframe
                frame_locator = self.page.frame_locator("iframe[name='bodyframe']")
                selector = "#datagrid-row-r1-2-0"
                await frame_locator.locator(selector).wait_for(state="visible", timeout=20000)
                await frame_locator.locator(selector).click()
                print("✅ Clicked 'Search Public Records'")
                return
            except Exception as e:
                print(f"⚠ click_search_public_records attempt {attempt} failed: {e}")
                if attempt == retries:
                    await self._dump_debug("click_search_public_records_failed")
                    raise
                # log frame URLs for debugging
                print("Frames at failure:")
                for f in self.page.frames:
                    print(" →", f.url)
                await asyncio.sleep(2)

    # -----------------------
    # Enter filing dates inside dynamic criteria frame
    # -----------------------
    async def enter_filing_dates(self, from_date: str = "01/01/2025", to_date: str = None, retries: int = 3):
        """Fill Filing Date From/To in the dynamically-loaded criteriaframe."""
        if to_date is None:
            to_date = datetime.today().strftime("%m/%d/%Y")

        print("⏳ Waiting a few seconds for dynamic frames to load...")
        await asyncio.sleep(3)

        for attempt in range(1, retries + 1):
            try:
                # Look for a frame whose URL contains 'dynCriteria.do'
                criteria_frame = None
                try:
                    criteria_frame = await self.wait_for_frame_by_url_fragment("dynCriteria.do", timeout=30)
                except PlaywrightTimeoutError:
                    # fallback: try blank.jsp or other known fragments
                    try:
                        criteria_frame = await self.wait_for_frame_by_url_fragment("blank.jsp", timeout=10)
                    except PlaywrightTimeoutError:
                        criteria_frame = None

                if not criteria_frame:
                    raise PlaywrightTimeoutError("Could not find criteria frame (dynCriteria.do or blank.jsp)")

                # ensure frame is loaded
                await criteria_frame.wait_for_load_state("domcontentloaded", timeout=15000)
                await asyncio.sleep(1)

                # Wait for date container
                el = await criteria_frame.wait_for_selector("#elemDateRange", timeout=15000)
                if not el:
                    raise PlaywrightTimeoutError("Date range container '#elemDateRange' not found")

                # locate FROM/TO inputs (IDs observed from your code)
                from_input = await criteria_frame.wait_for_selector("#_easyui_textbox_input7", timeout=10000)
                to_input = await criteria_frame.wait_for_selector("#_easyui_textbox_input8", timeout=10000)

                # Clear & type (use .fill where possible)
                await from_input.fill("")
                await from_input.type(from_date)
                print(f"✅ Entered FROM date: {from_date}")

                await to_input.fill("")
                await to_input.type(to_date)
                print(f"✅ Entered TO date: {to_date}")

                return

            except Exception as e:
                print(f"⚠ enter_filing_dates attempt {attempt} failed: {e}")
                if attempt == retries:
                    print("Frames (debug):")
                    for f in self.page.frames:
                        print(" ->", f.url)
                    await self._dump_debug("enter_filing_dates_failed")
                    raise
                await asyncio.sleep(2)




    # # -----------------------
    # # Click Search button inside dynamic search frame
    # # -----------------------
    async def click_search_button(self, retries: int = 3):
        """Click the 'Search Public Records' button by navigating through the iframe hierarchy."""
        for attempt in range(1, retries + 1):
            try:
                # First, wait for the main body frame
                body_frame = await self.wait_for_frame_by_name("bodyframe", timeout=30000)
                
                # Wait for the dynamic search frame inside the body frame
                dyn_search_frame = await self.wait_for_frame_by_name("dynSearchFrame", timeout=30000)
                
                # Wait for the frame to load the search criteria content
                await dyn_search_frame.wait_for_load_state("domcontentloaded", timeout=15000)
                
                # Wait for the search button to be available
                search_selector = "a[onclick*='executeSearchCommand'][onclick*='search']"
                await dyn_search_frame.wait_for_selector(search_selector, timeout=15000)
                
                # Click using JavaScript to ensure it works
                await dyn_search_frame.eval_on_selector(search_selector, "el => el.click()")
                
                print("✅ Clicked 'Search Public Records' button")
                return True
                
            except Exception as e:
                print(f"⚠ click_search_button attempt {attempt} failed: {e}")
                if attempt == retries:
                    print("❌ All attempts to click search button failed")
                    await self._dump_debug("search_button_failed")
                    return False
                await asyncio.sleep(2)

    async def wait_for_frame_by_name(self, name: str, timeout: float = 30000):
        """Wait for a frame with a specific name to be available."""
        start_time = time.time()
        while (time.time() - start_time) * 1000 < timeout:
            for frame in self.page.frames:
                if frame.name == name:
                    return frame
            await asyncio.sleep(0.1)
        raise PlaywrightTimeoutError(f"Frame with name '{name}' not found within {timeout}ms")





    # -----------------------
    # Example parse results hook (fill in selectors for results)
    # -----------------------
    # Add these new methods to the DelawareScraper class

    async def process_search_results(self):
        """Process all search results by iterating through pages and records."""
        all_data = []
        
        while True:
            # Process current page of results
            page_data = await self.process_current_page()
            all_data.extend(page_data)
            
            # Check if there's a next page
            has_next = await self.go_to_next_page()
            if not has_next:
                break
                
        return all_data

    async def process_current_page(self):
        """Process all records on the current results page."""
        page_data = []
        
        # Wait for results to load
        await asyncio.sleep(3)
        
        # Navigate to the results iframe
        result_list_frame = await self.wait_for_frame_by_name("resultListFrame", timeout=30000)
        
        # Get all record links
        record_links = await result_list_frame.query_selector_all("a.link[onclick*='loadRecord']")
        
        print(f"Found {len(record_links)} records on this page")
        
        # Process each record
        for i, link in enumerate(record_links):
            print(f"Processing record {i+1}/{len(record_links)}")
            
            # Click the record link
            await link.click()
            await asyncio.sleep(2)  # Wait for details to load
            
            # Extract data from details page
            record_data = await self.extract_record_data()
            if record_data:
                page_data.append(record_data)
                
            # Go back to results
            await self.go_back_to_results()
            await asyncio.sleep(1)
        
        return page_data

    async def extract_record_data(self):
        """Extract data from a record details page."""
        record_data = {}
        
        try:
            # Wait for document frame to load
            document_frame = await self.wait_for_frame_by_name("documentFrame", timeout=30000)
            
            # Wait for docInfoFrame
            doc_info_frame = await self.wait_for_frame_by_name("docInfoFrame", timeout=30000)
            
            # Click on Representatives tab
            reps_tab = await doc_info_frame.wait_for_selector("span.tabs-title:has-text('Representatives')", timeout=10000)
            await reps_tab.click()
            await asyncio.sleep(1)  # Wait for tab to load
            
            # Extract representative information
            rep_name = await doc_info_frame.query_selector("td:has-text('JULIE M BOWDERS')")
            if rep_name:
                record_data['representative_name'] = await rep_name.text_content()
                
            # Extract representative address
            address_elements = await doc_info_frame.query_selector_all("tr.evenrow td")
            address_parts = []
            for elem in address_elements:
                text = await elem.text_content()
                if text.strip() and 'JULIE M BOWDERS' not in text:
                    address_parts.append(text.strip())
            
            if address_parts:
                record_data['representative_address'] = ' '.join(address_parts)
            
            # Click on Decedent & Estate Info tab
            decedent_tab = await doc_info_frame.wait_for_selector("span.tabs-title:has-text('Decedent & Estate Info')", timeout=10000)
            await decedent_tab.click()
            await asyncio.sleep(1)  # Wait for tab to load
            
            # Extract filing date
            filing_date_elem = await doc_info_frame.query_selector("td:has-text('Filing Date:') + td")
            if filing_date_elem:
                record_data['filing_date'] = await filing_date_elem.text_content()
                
            # Extract decedent address
            address_elems = await doc_info_frame.query_selector_all("td:has-text('Address:'), td:has-text('City:'), td:has-text('State:'), td:has-text('Zip:')")
            address_data = {}
            
            for elem in address_elems:
                label = await elem.text_content()
                value_elem = await elem.evaluate_handle('elem => elem.nextElementSibling')
                if value_elem:
                    value = await value_elem.text_content()
                    address_data[label.replace(':', '').lower()] = value.strip()
            
            record_data['decedent_address'] = address_data
            
            return record_data
            
        except Exception as e:
            print(f"Error extracting record data: {e}")
            return None

    async def go_back_to_results(self):
        """Click the back to results button."""
        try:
            # Navigate to the bodyframe
            body_frame = await self.wait_for_frame_by_name("bodyframe", timeout=30000)
            
            # Find and click the back button
            back_button = await body_frame.wait_for_selector("a[onclick*='executeSearchNav'][onclick*='results']", timeout=10000)
            await back_button.click()
            await asyncio.sleep(2)  # Wait for results to load
            
            return True
        except Exception as e:
            print(f"Error going back to results: {e}")
            return False

    async def go_to_next_page(self):
        """Click the next page button if available."""
        try:
            # Navigate to the resultListFrame
            result_list_frame = await self.wait_for_frame_by_name("resultListFrame", timeout=30000)
            
            # Check if next page button exists
            next_button = await result_list_frame.query_selector("a[onclick*='navigateResults'][onclick*='next']")
            
            if next_button:
                await next_button.click()
                await asyncio.sleep(3)  # Wait for next page to load
                return True
            else:
                print("No more pages available")
                return False
                
        except Exception as e:
            print(f"Error going to next page: {e}")
            return False

    async def save_to_excel(self, data):
        """Save data to Excel with month-wise sheets."""
        try:
            import pandas as pd
            from openpyxl import Workbook
            
            # Create a new workbook
            wb = Workbook()
            
            # Group data by month
            monthly_data = {}
            for record in data:
                if 'filing_date' in record:
                    # Parse date and create month-year key
                    date_obj = datetime.strptime(record['filing_date'], '%m/%d/%Y')
                    month_key = date_obj.strftime('%Y-%m')
                    
                    if month_key not in monthly_data:
                        monthly_data[month_key] = []
                    monthly_data[month_key].append(record)
            
            # Create a sheet for each month
            for month, records in monthly_data.items():
                # Convert to DataFrame
                df = pd.DataFrame(records)
                
                # Create sheet
                if month in wb.sheetnames:
                    ws = wb[month]
                else:
                    ws = wb.create_sheet(month)
                
                # Write headers
                for col_idx, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                
                # Write data
                for row_idx, row in df.iterrows():
                    for col_idx, col_name in enumerate(df.columns, 1):
                        ws.cell(row=row_idx+2, column=col_idx, value=row[col_name])
            
            # Remove default sheet if empty
            if 'Sheet' in wb.sheetnames and wb['Sheet'].max_row == 1:
                del wb['Sheet']
            
            # Save file
            filename = OUT_DIR / f"delaware_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            print(f"✅ Data saved to {filename}")
            
        except ImportError:
            print("⚠ openpyxl or pandas not installed. Saving as JSON instead.")
            filename = OUT_DIR / f"delaware_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"✅ Data saved to {filename}")



async def main():
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=False)
        context = await browser.new_context()
        page = await context.new_page()

        scraper = DelawareScraper(page, browser=browser, context=context)
        try:
            await scraper.goto_login()
            await scraper.accept_terms()
            await scraper.click_search_public_records()
            await scraper.enter_filing_dates()
            await scraper.click_search_button()
            #wait for results to load
            await asyncio.sleep(10)
            # Process all search results
            all_data = await scraper.process_search_results()
            
            # Save to Excel
            await scraper.save_to_excel(all_data)
            
            print("✅ Scrape finished successfully")
            
        except Exception:
            print("❌ Scraper failed with exception:")
            traceback.print_exc()
            try:
                await scraper._dump_debug("fatal_error")
            except Exception as e:
                print("debug dump also failed:", e)
            raise
        finally:
            await browser.close()

if __name__ == "__main__":
    asyncio.run(main())

