# scrape_delaware.py
import asyncio
import json
import os
import traceback
from datetime import datetime, timedelta
from pathlib import Path
import time
from collections import defaultdict
from urllib.parse import urlparse, parse_qs

from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Google Sheets imports
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import datetime as dt

# -----------------------------
# Google Sheets Configuration
# -----------------------------
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
SPREADSHEET_ID = os.environ.get("SPREADSHEET_ID")

# -----------------------------
# Google Sheets Helpers
# -----------------------------
def load_service_account_info():
    file_env = os.environ.get("GOOGLE_CREDENTIALS_FILE")
    if file_env:
        if os.path.exists(file_env):
            with open(file_env, "r", encoding="utf-8") as fh:
                return json.load(fh)
        raise ValueError(f"GOOGLE_CREDENTIALS_FILE set but not found: {file_env}")

    creds_raw = os.environ.get("GOOGLE_CREDENTIALS")
    if not creds_raw:
        raise ValueError("GOOGLE_CREDENTIALS or GOOGLE_CREDENTIALS_FILE is required.")

    txt = creds_raw.strip()
    if txt.startswith("{"):
        return json.loads(txt)

    if os.path.exists(creds_raw):
        with open(creds_raw, "r", encoding="utf-8") as fh:
            return json.load(fh)

    raise ValueError("GOOGLE_CREDENTIALS is neither valid JSON nor an existing file path.")

def sheets_client():
    info = load_service_account_info()
    creds = service_account.Credentials.from_service_account_info(info, scopes=SCOPES)
    return build("sheets", "v4", credentials=creds).spreadsheets()

def get_last_scraped_date(svc, spreadsheet_id, sheet_name="RawData"):
    """Get the most recent filing date from Google Sheets."""
    try:
        # Column B = Filing Date, start from row 2 (skip header)
        res = svc.values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'!B2:B"
        ).execute()
        vals = [r[0] for r in res.get("values", []) if r]

        if not vals:
            return None

        parsed = []
        for d in vals:
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
                try:
                    parsed.append(dt.datetime.strptime(d, fmt).date())
                    break
                except ValueError:
                    continue
        return max(parsed) if parsed else None

    except Exception as e:
        print(f"⚠ Error reading last date from {sheet_name}: {e}")
        return None

def append_to_google_sheets(svc, spreadsheet_id, records, sheet_name="RawData"):
    """Append new records to Google Sheets."""
    if not records:
        print("📭 No new records to append to Google Sheets")
        return

    # Prepare data for Google Sheets
    values = []
    for record in records:
        values.append([
            record.get("case_file_no", ""),
            record.get("filing_date", ""),
            record.get("caseFileNum", ""),
            record.get("caseFileId", ""),
            record.get("decedent_address", ""),
            record.get("representative_name", ""),
            record.get("representative_address", "")
        ])

    try:
        # Ensure sheet exists
        ensure_sheet_exists(svc, spreadsheet_id, sheet_name)
        
        # Append data
        result = svc.values().append(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'!A:G",
            valueInputOption="USER_ENTERED",
            body={"values": values}
        ).execute()
        
        print(f"✅ Appended {len(records)} records to Google Sheets ({sheet_name})")
        return result
    except Exception as e:
        print(f"❌ Error appending to Google Sheets: {e}")
        return None

def ensure_sheet_exists(svc, spreadsheet_id, sheet_name):
    """Create a sheet if missing with proper headers."""
    try:
        # Get existing sheets
        spreadsheet = svc.get(spreadsheetId=spreadsheet_id).execute()
        sheets = spreadsheet.get('sheets', [])
        existing_titles = [sheet['properties']['title'] for sheet in sheets]
        
        if sheet_name not in existing_titles:
            # Create new sheet
            body = {
                "requests": [{
                    "addSheet": {
                        "properties": {
                            "title": sheet_name
                        }
                    }
                }]
            }
            svc.batchUpdate(spreadsheetId=spreadsheet_id, body=body).execute()
            
            # Add headers
            headers = [
                "Case File No", "Filing Date", "Case File Num", "Case File ID", 
                "Decedent Address", "Representative Name", "Representative Address"
            ]
            svc.values().update(
                spreadsheetId=spreadsheet_id,
                range=f"'{sheet_name}'!A1:G1",
                valueInputOption="USER_ENTERED",
                body={"values": [headers]}
            ).execute()
            
            print(f"✅ Created new sheet: {sheet_name}")
            
    except Exception as e:
        print(f"⚠ Error ensuring sheet exists: {e}")

def normalize_date(datestr):
    """Convert various date formats to MM/DD/YYYY for consistent scraping."""
    if not datestr:
        return ""
    
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            dt_obj = dt.datetime.strptime(datestr, fmt)
            return dt_obj.strftime("%m/%d/%Y")
        except ValueError:
            continue
    return datestr  # fallback unchanged

OUT_DIR = Path("out")
OUT_DIR.mkdir(exist_ok=True)

class DelawareScraper:
    def __init__(self, page, browser=None, context=None,
                 base_url: str = "https://delcorowonlineservices.co.delaware.pa.us/countyweb/loginDisplay.action?countyname=DelawarePA"):
        self.page = page
        self.browser = browser
        self.context = context
        self.base_url = base_url

    # === FRAME LOCATOR METHODS ===
    def _res_list_loc(self):
        return (
            self.page
            .frame_locator("iframe[name='bodyframe']")
            .frame_locator("iframe[name='resultFrame']")
            .frame_locator("iframe[name='resultListFrame']")
        )

    def _res_subnav_middle_loc(self):
        return (
            self.page
            .frame_locator("iframe[name='bodyframe']")
            .frame_locator("iframe[name='resultFrame']")
            .frame_locator("iframe[src*='navbar.do?page=search.resultNav.middle']")
        )

    def _doc_loc(self):
        return (
            self.page
            .frame_locator("iframe[name='bodyframe']")
            .frame_locator("iframe[name='documentFrame']")
            .frame_locator("iframe[name='docInfoFrame']")
        )

    def _tabs_loc(self):
        return self._doc_loc().frame_locator("iframe[name='tabs']")

    async def ensure_decedent_tab(self):
        """Ensure the decedent panel is visible."""
        try:
            if await self._doc_loc().locator("text=Estate Info").first.is_visible():
                return
        except:
            pass
        
        tabs = self._tabs_loc()
        for label in ["Decedent & Estate Info", "Decedent", "Estate Info"]:
            try:
                await tabs.locator(f"span.tabs-title:has-text('{label}')").first.click(timeout=2000)
                await asyncio.sleep(0.6)
                return
            except:
                continue

    async def _dump_debug(self, name_prefix: str):
        """Save HTML for debugging."""
        ts = datetime.now().strftime("%Y%m%dT%H%M%SZ")
        html = OUT_DIR / f"{name_prefix}-{ts}.html"
        try:
            html_content = await self.page.content()
            html.write_text(html_content, encoding="utf-8")
            print(f"Debug HTML dumped: {html}")
        except Exception as e:
            print("Failed to save HTML:", e)

    async def wait_for_frame_by_url_fragment(self, url_fragment: str, timeout: int = 60):
        """Poll the page.frames until a frame whose URL contains url_fragment appears."""
        print(f"Waiting for frame with url containing '{url_fragment}' (timeout {timeout}s)...")
        for i in range(timeout):
            for f in self.page.frames:
                if f.url and url_fragment in f.url:
                    print(f"Found frame with url {f.url}")
                    return f
            await asyncio.sleep(1)
        raise PlaywrightTimeoutError(f"Frame with url fragment '{url_fragment}' not found within {timeout}s")

    # -----------------------
    # Login / navigation 
    # -----------------------
    async def goto_login(self, retries: int = 3):
        """Go to login page and click 'Login as Guest'."""
        for attempt in range(1, retries + 1):
            try:
                await self.page.goto(self.base_url, timeout=60000)
                await self.page.wait_for_timeout(1500)
                possible_selectors = [
                    "input[value=' Login as Guest ']",
                    "input[value='Login as Guest']",
                    "input[type='button'][value*='Guest']",
                    "text='Login as Guest'"
                ]
                clicked = False
                for sel in possible_selectors:
                    try:
                        await self.page.wait_for_selector(sel, timeout=5000)
                        await self.page.locator(sel).click()
                        clicked = True
                        break
                    except PlaywrightTimeoutError:
                        continue
                if not clicked:
                    raise PlaywrightTimeoutError("Could not find 'Login as Guest' button with known selectors")
                await self.page.wait_for_url("**/main.jsp?countyname=DelawarePA", timeout=60000)
                print("✅ Logged in as Guest")
                return
            except Exception as e:
                print(f"⚠ goto_login attempt {attempt} failed: {e}")
                if attempt == retries:
                    await self._dump_debug("goto_login_failed")
                    raise
                await asyncio.sleep(3)

    async def accept_terms(self, retries: int = 3):
        """Handle iframe and click Accept button using frame_locator."""
        for attempt in range(1, retries + 1):
            try:
                await asyncio.sleep(1)
                iframe_selectors = [
                    "iframe[name='bodyframe']",
                    "iframe#bodyframe",
                    "iframe[src*='blank.jsp']",
                    "iframe",
                ]
                clicked = False
                for ifsel in iframe_selectors:
                    try:
                        frame_locator = self.page.frame_locator(ifsel)
                        accept = frame_locator.locator("#accept")
                        await accept.wait_for(state="visible", timeout=10000)
                        await accept.click()
                        clicked = True
                        break
                    except PlaywrightTimeoutError:
                        continue

                if not clicked:
                    for f in self.page.frames:
                        try:
                            handle = await f.query_selector("#accept")
                            if handle:
                                await handle.click()
                                clicked = True
                                break
                        except Exception:
                            continue

                if not clicked:
                    raise PlaywrightTimeoutError("Could not locate an Accept button in known frame locations")

                print(f"✅ Accepted terms (attempt {attempt})")
                return

            except Exception as e:
                print(f"⚠ accept_terms attempt {attempt} failed: {e}")
                if attempt == retries:
                    print("Available frames (debug):")
                    for f in self.page.frames:
                        print(" -", f.url)
                    await self._dump_debug("accept_terms_failed")
                    raise
                await asyncio.sleep(2)

    async def click_search_public_records(self, retries: int = 3):
        """Click the row that triggers Search Public Records."""
        for attempt in range(1, retries + 1):
            try:
                await self.page.wait_for_timeout(1000)
                frame_locator = self.page.frame_locator("iframe[name='bodyframe']")
                selector = "#datagrid-row-r1-2-0"
                await frame_locator.locator(selector).wait_for(state="visible", timeout=20000)
                await frame_locator.locator(selector).click()
                print("✅ Clicked 'Search Public Records'")
                return
            except Exception as e:
                print(f"⚠ click_search_public_records attempt {attempt} failed: {e}")
                if attempt == retries:
                    await self._dump_debug("click_search_public_records_failed")
                    raise
                print("Frames at failure:")
                for f in self.page.frames:
                    print(" →", f.url)
                await asyncio.sleep(2)

    async def enter_filing_dates(self, from_date: str = "01/01/2025", to_date: str = None, retries: int = 3):
        """Fill Filing Date From/To in the dynamically-loaded criteriaframe."""
        if to_date is None:
            to_date = datetime.today().strftime("%m/%d/%Y")

        print(f"⏳ Setting date range: {from_date} to {to_date}")
        await asyncio.sleep(3)

        for attempt in range(1, retries + 1):
            try:
                criteria_frame = None
                try:
                    criteria_frame = await self.wait_for_frame_by_url_fragment("dynCriteria.do", timeout=30)
                except PlaywrightTimeoutError:
                    try:
                        criteria_frame = await self.wait_for_frame_by_url_fragment("blank.jsp", timeout=10)
                    except PlaywrightTimeoutError:
                        criteria_frame = None

                if not criteria_frame:
                    raise PlaywrightTimeoutError("Could not find criteria frame (dynCriteria.do or blank.jsp)")

                await criteria_frame.wait_for_load_state("domcontentloaded", timeout=15000)
                await asyncio.sleep(1)

                el = await criteria_frame.wait_for_selector("#elemDateRange", timeout=15000)
                if not el:
                    raise PlaywrightTimeoutError("Date range container '#elemDateRange' not found")

                from_input = await criteria_frame.wait_for_selector("#_easyui_textbox_input7", timeout=10000)
                to_input = await criteria_frame.wait_for_selector("#_easyui_textbox_input8", timeout=10000)

                await from_input.fill("")
                await from_input.type(from_date)
                print(f"✅ Entered FROM date: {from_date}")

                await to_input.fill("")
                await to_input.type(to_date)
                print(f"✅ Entered TO date: {to_date}")

                return

            except Exception as e:
                print(f"⚠ enter_filing_dates attempt {attempt} failed: {e}")
                if attempt == retries:
                    print("Frames (debug):")
                    for f in self.page.frames:
                        print(" ->", f.url)
                    await self._dump_debug("enter_filing_dates_failed")
                    raise
                await asyncio.sleep(2)

    async def click_search_button(self, retries: int = 3):
        """Click the 'Search Public Records' button."""
        for attempt in range(1, retries + 1):
            try:
                body_frame = await self.wait_for_frame_by_name("bodyframe", timeout=30000)
                dyn_search_frame = await self.wait_for_frame_by_name("dynSearchFrame", timeout=30000, parent_frame=body_frame)
                await dyn_search_frame.wait_for_load_state("domcontentloaded", timeout=15000)
                search_selector = "a[onclick*='executeSearchCommand'][onclick*='search']"
                await dyn_search_frame.wait_for_selector(search_selector, timeout=15000)
                await dyn_search_frame.eval_on_selector(search_selector, "el => el.click()")
                print("✅ Clicked 'Search Public Records' button")
                return True
            except Exception as e:
                print(f"⚠ click_search_button attempt {attempt} failed: {e}")
                if attempt == retries:
                    print("❌ All attempts to click search button failed")
                    await self._dump_debug("search_button_failed")
                    return False
                await asyncio.sleep(2)

    async def wait_for_frame_by_name(self, name: str, timeout: float = 30000, parent_frame=None):
        """Wait for a frame with a specific name to be available."""
        start_time = time.time()
        while (time.time() - start_time) * 1000 < timeout:
            frames = parent_frame.child_frames if parent_frame else self.page.frames
            for frame in frames:
                if frame.name == name:
                    return frame
            await asyncio.sleep(0.1)
        raise PlaywrightTimeoutError(f"Frame with name '{name}' not found within {timeout}ms")

    # === EXTRACTION METHODS ===
    async def extract_decedent_info_atomic(self):
        """Extract decedent info using stable frame locators."""
        loc = self._doc_loc()

        # Ensure correct tab is active
        await self.ensure_decedent_tab()

        # Extract filing date
        filing_date = ""
        for _ in range(30):
            try:
                fd_cell = loc.locator("#fieldFILING_DATEspan").locator("xpath=ancestor::tr/td[3]").first
                txt = await fd_cell.text_content()
                if txt and txt.strip():
                    filing_date = txt.strip()
                    break
            except:
                pass
            await asyncio.sleep(0.5)

        # Extract case file number
        case_file_no = ""
        try:
            cf_cell = loc.locator("#fieldCASENUMBERspan").locator("xpath=ancestor::tr/td[3]").first
            case_file_no = (await cf_cell.text_content() or "").strip()
        except:
            pass

        # Extract address components
        addr = city = state = zipc = ""

        try:
            addr_cell = loc.locator("#fcaddrCORESPONDENT_ADDRESSspan").locator("xpath=ancestor::tr/td[3]").first
            addr = (await addr_cell.text_content() or "").strip()
        except:
            pass

        try:
            city_cell = loc.locator("#fccityCORESPONDENT_ADDRESSspan").locator("xpath=ancestor::tr/td[3]").first
            city = (await city_cell.text_content() or "").strip()
        except:
            pass

        try:
            # state and zip live in nested table
            st_row = loc.locator("#fcstateCORESPONDENT_ADDRESSspan").locator("xpath=ancestor::tr").first
            st_cell = st_row.locator("td").nth(2).locator("table td").nth(0)
            zp_cell = st_row.locator("td").nth(2).locator("table td").nth(2)
            state = ((await st_cell.text_content()) or "").strip()
            zipc  = ((await zp_cell.text_content()) or "").strip()
        except:
            pass

        parts = [p for p in [addr, city, state, zipc] if p]
        decedent_address = ", ".join(parts) if parts else ""

        return {
            "case_file_no": case_file_no,
            "filing_date": filing_date,
            "decedent_address": decedent_address
        }

    async def extract_representatives_atomic(self):
        """Extract representatives using stable frame locators."""
        loc = self._doc_loc()
        
        # Wait for representatives content
        try:
            await loc.locator("text=Personal Representative").first.wait_for(timeout=4000)
        except:
            await loc.locator("tr.evenrow, tr.oddrow").first.wait_for(timeout=6000)

        # Extract representative data
        rows = await loc.locator("tr.evenrow, tr.oddrow").all_text_contents()
        reps = []
        current = {"name": "", "address": ""}

        def looks_like_address(t: str) -> bool:
            u = t.upper()
            return any(ch.isdigit() for ch in t) or any(k in u for k in [
                "AVE","ST","STREET","AVENUE","ROAD","RD","LANE","LN","DR","DRIVE",
                "APT","SUITE","PO BOX","BLVD","COURT","CT"
            ])

        for raw in rows:
            t = raw.strip()
            if not t:
                continue
            if not looks_like_address(t) and len(t) < 120:
                if current["name"]:
                    reps.append({
                        "representative_name": current["name"], 
                        "representative_address": current["address"]
                    })
                    current = {"name": "", "address": ""}
                current["name"] = t
            else:
                current["address"] = f"{current['address']} {t}".strip() if current["address"] else t

        if current["name"]:
            reps.append({
                "representative_name": current["name"], 
                "representative_address": current["address"]
            })

        return reps

    async def safe_click_tab(self, tab_text, retries=3):
        """Safely click a tab by text using frame locators."""
        tabs = self._tabs_loc()
        for attempt in range(retries):
            try:
                for sel in [
                    f"span.tabs-title:has-text('{tab_text}')",
                    f"li:has-text('{tab_text}') span.tabs-inner",
                    f"li:has-text('{tab_text}')",
                    f"span:has-text('{tab_text}')",
                ]:
                    try:
                        await tabs.locator(sel).first.click(timeout=4000)
                        return True
                    except:
                        continue

                # Fallback: JS click
                try:
                    await tabs.evaluate(f"""
                        () => {{
                            const spans = Array.from(document.querySelectorAll('span.tabs-title'));
                            const t = spans.find(x => x.textContent.trim().includes('{tab_text}'));
                            if (t) {{ t.click(); return true; }}
                            const li = Array.from(document.querySelectorAll('li')).find(x => x.textContent.includes('{tab_text}'));
                            if (li) {{ li.click(); return true; }}
                            return false;
                        }}
                    """)
                    return True
                except:
                    pass
            except:
                if attempt < retries - 1:
                    await asyncio.sleep(0.8)
        return False

    async def click_result_link_by_index(self, index: int) -> bool:
        """Click row N on the current page using stable frame locators."""
        res = self._res_list_loc()

        # Try primary selector
        a = res.locator(f"a.link#inst{index}").first
        if await a.count():
            try:
                await a.click(timeout=6000)
                return True
            except:
                pass

        # Fallback: Nth clickable link
        all_links = res.locator("a.link[onclick*='loadRecord']")
        if await all_links.count() > index:
            try:
                await all_links.nth(index).click(timeout=6000)
                return True
            except:
                pass

        return False

    async def click_back_to_results(self, retries=5):
        """Return to results page using working navigation."""
        for attempt in range(retries):
            try:
                await asyncio.sleep(0.8)
                bodyframe = await self.wait_for_frame_by_name("bodyframe", 10000)

                resnavframe = None
                for _ in range(10):
                    for f in bodyframe.child_frames:
                        if f.name == "resnavframe":
                            resnavframe = f
                            break
                    if resnavframe:
                        break
                    await asyncio.sleep(0.3)

                if not resnavframe:
                    resnavframe = await self.wait_for_frame_by_url_fragment("navbar.do?page=search.details", 10)

                await resnavframe.wait_for_load_state("domcontentloaded", timeout=10000)
                await asyncio.sleep(0.5)

                clicked = False
                for sel in [
                    "text='Back to Results'",
                    "a[onclick*='executeSearchNav'][onclick*='results']",
                    "img[alt='Back to Results']",
                ]:
                    try:
                        await resnavframe.click(sel, timeout=3000)
                        clicked = True
                        break
                    except:
                        pass

                if not clicked:
                    try:
                        await resnavframe.evaluate("""
                            () => {
                                const elements = Array.from(document.querySelectorAll('a, img'));
                                const backElement = elements.find(el =>
                                    el.textContent.includes('Back to Results') ||
                                    el.alt === 'Back to Results' ||
                                    (el.onclick && el.onclick.toString().includes('results'))
                                );
                                if (backElement) {
                                    if (backElement.onclick) backElement.onclick();
                                    else if (backElement.parentElement && backElement.parentElement.onclick) backElement.parentElement.onclick();
                                    else backElement.click();
                                    return true;
                                }
                                return false;
                            }
                        """)
                        clicked = True
                    except:
                        pass

                if clicked:
                    await asyncio.sleep(1.2)
                    try:
                        await self.wait_for_frame_by_name("bodyframe", 5000)
                        await self.wait_for_frame_by_url_fragment("SearchResultsView.jsp", 5)
                        print("✅ Returned to results page")
                        return True
                    except:
                        print(f"Back button clicked but results verification failed, attempt {attempt + 1}")
                        continue
                else:
                    print(f"Could not find back button, attempt {attempt + 1}")

            except Exception as e:
                print(f"Back to results attempt {attempt + 1} failed: {e}")
                if attempt < retries - 1:
                    await asyncio.sleep(1.2)
                continue

        print("❌ Failed to return to results after all attempts")
        return False

    async def scrape_all_records_via_next_button(self, from_date: str = "01/01/2025", to_date: str = None):
        """Modified to accept custom date range for incremental scraping."""
        if to_date is None:
            to_date = datetime.today().strftime("%m/%d/%Y")

        all_records = []
        page_index = 1
        max_pages = 72  # Adjust as needed

        print(f"🎯 Starting incremental record scraping from {from_date} to {to_date}...")

        while page_index <= max_pages:
            print(f"\n{'='*50}")
            print(f"📄 PROCESSING PAGE {page_index}")
            print(f"{'='*50}")

            # Wait for results list to be ready
            try:
                await self._res_list_loc().locator("a.link#inst0, a.link[onclick*='loadRecord']").first.wait_for(timeout=15000)
            except:
                print(f"❌ No results found on page {page_index}. Stopping.")
                break

            processed_this_page = 0
            consecutive_misses = 0

            # Process up to 40 records per page
            for row_idx in range(40):
                print(f"  📝 Processing record {row_idx + 1} of 40 on page {page_index}")

                # Click the record link
                success = await self.click_result_link_by_index(row_idx)
                if not success:
                    print(f"  ❌ Could not click record {row_idx + 1}")
                    consecutive_misses += 1
                    if consecutive_misses >= 3:
                        print("  ⚠ Several consecutive misses, assuming end of page")
                        break
                    continue

                # Wait for document frame to load
                try:
                    await self._doc_loc().locator("body").first.wait_for(state="attached", timeout=20000)
                    await asyncio.sleep(1.0)
                except Exception as e:
                    print(f"  ❌ Failed to load details for record {row_idx + 1}: {e}")
                    consecutive_misses += 1
                    continue

                # Extract decedent info
                dec_info = {}
                try:
                    dec_info = await self.extract_decedent_info_atomic()
                    print(f"  ✅ Decedent info extracted")
                except Exception as e:
                    print(f"  ⚠ Decedent extraction failed: {e}")

                # Extract representatives
                reps = []
                try:
                    clicked = await self.safe_click_tab("Representatives", retries=2)
                    if clicked:
                        await asyncio.sleep(1.0)
                        reps = await self.extract_representatives_atomic()
                        print(f"  ✅ {len(reps)} representatives extracted")
                except Exception as e:
                    print(f"  ⚠ Representatives extraction failed: {e}")

                # Get case metadata from URL
                case_meta = {}
                try:
                    for f in self.page.frames:
                        if f.url and "DocumentInfoView.jsp" in f.url and "caseFileId=" in f.url:
                            qs = parse_qs(urlparse(f.url).query)
                            case_meta = {
                                "caseFileId": (qs.get("caseFileId") or [""])[0],
                                "caseFileNum": (qs.get("caseFileNum") or [""])[0],
                            }
                            break
                except Exception as e:
                    print(f"  ⚠ Case metadata error: {e}")

                # Combine data
                base_record = {
                    "case_file_no": dec_info.get("case_file_no", ""),
                    "filing_date": dec_info.get("filing_date", ""),
                    "decedent_address": dec_info.get("decedent_address", ""),
                    **case_meta,
                }

                if reps:
                    record_data = [{**base_record, **rep} for rep in reps]
                else:
                    record_data = [{**base_record, "representative_name": "", "representative_address": ""}]

                all_records.extend(record_data)
                processed_this_page += len(record_data)
                consecutive_misses = 0

                print(f"  ✅ Record {row_idx + 1} processed: {len(record_data)} entries")

                # Return to results page
                back_success = await self.click_back_to_results()
                if not back_success:
                    print("  ❌ Failed to return to results, stopping page processing")
                    break

                await asyncio.sleep(0.5)

            print(f"✅ Page {page_index} complete: {processed_this_page} records extracted")

            # Navigate to next page
            if page_index < max_pages:
                print(f"🔄 Navigating to page {page_index + 1}...")
                next_success = await self.goto_results_page(page_index + 1)
                if not next_success:
                    print("❌ Failed to navigate to next page, stopping")
                    break
                page_index += 1
                await asyncio.sleep(1.0)
            else:
                break

        print(f"\n{'='*50}")
        print(f"🎉 SCRAPING COMPLETE")
        print(f"📊 {len(all_records)} total records extracted")
        print(f"{'='*50}")

        return all_records

    async def goto_results_page(self, page_number: int, wait_timeout: int = 20000) -> bool:
        """Robust navigation to a specific results page."""
        print(f"🔎 goto_results_page -> target {page_number}")
        # First try the known subnav locator (fast path)
        try:
            subnav = self._res_subnav_middle_loc()
            input_locator = subnav.locator("input[name='pageNumber']").first
            await input_locator.wait_for(state="attached", timeout=6000)
            await input_locator.fill(str(page_number))

            # Try clicking a Go link if present
            go_link = subnav.locator("a[onclick*='goToResultPage'], a[onclick*='goToResultPage()']").first
            if await go_link.count():
                try:
                    await go_link.click()
                except Exception:
                    # fallback to evaluating JS on the frame
                    await subnav.evaluate("() => { if (window.goToResultPage) goToResultPage(); }")
            else:
                # Try calling the function directly in that frame
                await subnav.evaluate("() => { if (window.goToResultPage) goToResultPage(); }")

            # wait for results list to refresh
            res = self._res_list_loc()
            await res.locator("a.link#inst0, a.link[onclick*='loadRecord']").first.wait_for(timeout=wait_timeout)
            await asyncio.sleep(0.6)
            print(f"✅ Navigated to page {page_number} (fast path)")
            return True

        except Exception as fast_err:
            print(f"⚠ fast path for goto_results_page failed: {fast_err}")

        # Fallback: search frames for an input[name='pageNumber']
        try:
            input_frame = None
            input_element = None

            for f in self.page.frames:
                try:
                    handles = await f.query_selector_all("input[name='pageNumber']")
                    if handles:
                        input_frame = f
                        input_element = handles[0]
                        break
                except Exception:
                    continue

            if not input_element:
                print("❌ Could not find pageNumber input in any frame (fallback)")
                return False

            # Fill the input using the frame handle
            await input_element.fill(str(page_number))

            # Try to call the JS navigation function in that frame
            try:
                result = await input_frame.evaluate(
                    """() => {
                        try {
                            if (typeof goToResultPage === 'function') { goToResultPage(); return true; }
                            if (window.goToResultPage) { window.goToResultPage(); return true; }
                            // try clicking any nearby 'Go' button if present
                            const goBtn = Array.from(document.querySelectorAll('a, input, button')).find(el =>
                                (el.getAttribute && (el.getAttribute('onclick') || '').includes('goToResultPage')) ||
                                (el.textContent && /go\s*to\s*result/i.test(el.textContent))
                            );
                            if (goBtn) { goBtn.click(); return true; }
                            return false;
                        } catch (e) {
                            return false;
                        }
                    }"""
                )
                if not result:
                    print("⚠ Fallback JS navigation did not report success; continuing to wait for result marker")

            except Exception as eval_err:
                print(f"⚠ Evaluating JS in fallback frame failed: {eval_err}")

            # Wait for results list to refresh
            try:
                res = self._res_list_loc()
                await res.locator("a.link#inst0, a.link[onclick*='loadRecord']").first.wait_for(timeout=wait_timeout)
                await asyncio.sleep(0.6)
                print(f"✅ Navigated to page {page_number} (fallback)")
                return True
            except Exception as wait_err:
                print(f"❌ Waiting for results after fallback navigation failed: {wait_err}")
                return False

        except Exception as e:
            print(f"❌ goto_results_page unexpected failure: {e}")
            return False

    def write_monthwise_xlsx(self, records, out_path):
        """Write records to XLSX with month-wise sheets."""
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        by_month = defaultdict(list)
        for r in records:
            fd = (r.get("filing_date") or "").strip()
            month_key = "Unknown"
            try:
                dt = datetime.strptime(fd, "%m/%d/%Y")
                month_key = dt.strftime("%Y-%m")
            except Exception:
                try:
                    dt = datetime.strptime(fd, "%m/%d/%y")
                    month_key = dt.strftime("%Y-%m")
                except Exception:
                    pass
            by_month[month_key].append(r)

        headers = ["case_file_no", "filing_date", "caseFileNum", "caseFileId", "decedent_address", "representative_name", "representative_address"]

        for month in sorted(by_month.keys()):
            ws = wb.create_sheet(title=month[:31])
            ws.append(headers)
            for r in by_month[month]:
                ws.append([r.get(h, "") for h in headers])

            # Auto-adjust column widths
            for col_idx, h in enumerate(headers, start=1):
                max_len = max([len(str(h))] + [len(str(ws.cell(row=i, column=col_idx).value or "")) for i in range(2, ws.max_row + 1)])
                ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max_len + 2)

        wb.save(out_path)
        print(f"✅ XLSX written: {out_path}")


async def run_incremental_scrape_and_export(scraper, from_date: str, to_date: str):
    """
    Run incremental scraping process and export results.
    Only keep records that have representative_name and representative_address.
    """
    print(f"🚀 Starting incremental scraping from {from_date} to {to_date}...")
    all_records = await scraper.scrape_all_records_via_next_button(from_date, to_date)

    # ✅ Filter records to only keep those with representative info
    filtered_records = [
        r for r in all_records
        if r.get("representative_name") and r.get("representative_address")
    ]

    return filtered_records


async def main():
    """Main function with incremental scraping and Google Sheets integration."""
    # Check required environment variables
    if not SPREADSHEET_ID:
        print("❌ SPREADSHEET_ID environment variable is required")
        return
    
    # Initialize Google Sheets client
    try:
        svc = sheets_client()
        print("✅ Google Sheets client initialized")
    except Exception as e:
        print(f"❌ Failed to initialize Google Sheets client: {e}")
        return

    # Get last scraped date from Google Sheets
    last_date = get_last_scraped_date(svc, SPREADSHEET_ID)
    
    if last_date:
        # Start from day after last recorded date
        from_date = (last_date + timedelta(days=1)).strftime("%m/%d/%Y")
        print(f"📅 Resuming from last recorded date: {last_date} -> {from_date}")
    else:
        # Default start date if no previous data
        from_date = "01/01/2025"
        print("📅 No previous data found, starting from default date: 01/01/2025")
    
    # End date is today
    to_date = datetime.today().strftime("%m/%d/%Y")
    
    if last_date and last_date >= datetime.today().date():
        print("✅ No new records to scrape - already up to date")
        return

    async with async_playwright() as pw:
        if os.environ.get("GITHUB_ACTIONS") == "true":
            # 👉 Running in GitHub Actions: no profile dir
            print("Running in GitHub Actions: using fresh Chromium context")
            browser = await pw.chromium.launch(
                headless=True,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--disable-dev-shm-usage",
                    "--no-sandbox",
                ],
            )
            context = await browser.new_context()
        else:
            # 👉 Local run: keep using persistent profile
            profile_dir = Path(os.environ.get("CHROME_PROFILE_DIR", "chrome_profile")).resolve()
            profile_dir.mkdir(exist_ok=True)
            print(f"Using Chrome profile: {profile_dir}")
            context = await pw.chromium.launch_persistent_context(
                user_data_dir=str(profile_dir),
                channel="chrome",
                headless=False,
                ignore_https_errors=True,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--no-first-run",
                    "--no-default-browser-check",
                    "--disable-dev-shm-usage",
                    "--no-sandbox",
                ],
            )

        page = context.pages[0] if context.pages else await context.new_page()
        scraper = DelawareScraper(page, browser=None, context=context)
        
        try:
            # Execute the scraping workflow
            await scraper.goto_login()
            await scraper.accept_terms()
            await scraper.click_search_public_records()
            await scraper.enter_filing_dates(from_date=from_date, to_date=to_date)
            await scraper.click_search_button()
            
            # Wait for results to load
            await asyncio.sleep(10)
            
            # Run incremental scraping and export
            new_records = await run_incremental_scrape_and_export(scraper, from_date, to_date)
            
            # Append to Google Sheets
            if new_records:
                append_to_google_sheets(svc, SPREADSHEET_ID, new_records)
                print(f"🎉 Successfully processed {len(new_records)} new records")
            else:
                print("📭 No new records found with representative information")
            
        except Exception as e:
            print("❌ Scraper failed with exception:")
            traceback.print_exc()
            try:
                await scraper._dump_debug("fatal_error")
            except Exception as debug_error:
                print("Debug dump also failed:", debug_error)
            raise
        finally:
            await context.close()


if __name__ == "__main__":
    asyncio.run(main())